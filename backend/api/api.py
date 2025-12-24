from fastapi import FastAPI, UploadFile, File, Form, Request, BackgroundTasks, Header
# import pandas as pd  # pandas 없이 작동하도록 주석 처리
import io
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from bs4 import BeautifulSoup
from typing import List, Dict, Any
import requests

import os
import re
import shutil
import logging
import json
import time
from datetime import datetime


from ..utils.config import STATIC_DIR, DOCS_DIR, LOG_DIR, CHROMA_DIR
from ..data.database import (
    get_vectordb,
    index_document,
    get_indexed_files,
    remove_document,
    reset_vectordb,
    save_feedback_to_file,
    get_feedback_vectordb,
    index_json_data
)
# semantic_search 모듈 제거됨
from ..utils.slack import clean_mention, post_slack_reply, handle_slack_message
from ..utils.utils import format_sources

# qa_utils 모듈 제거됨
from ..services.category_classifier import auto_classify
from slack_sdk.web.async_client import AsyncWebClient
from ..services.effort_estimation import EffortEstimation, effort_manager
from ..services.effort_qa import run_effort_qa_chain, run_effort_qa_with_feedback, get_effort_statistics, search_similar_features
from ..data.database import get_vectordb, index_document, index_json_data, index_json_data_incremental
from ..services.jira_integration import create_jira_integration
from ..services.mock_qa import mock_qa_response, mock_effort_qa_response
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..', '..'))
from data.prompts import intent_prompt_manager
# from apscheduler.schedulers.background import BackgroundScheduler  # SSL 문제로 임시 비활성화

# Configure logging
from ..utils.config import LOG_DIR

# 로그 디렉토리 확인
os.makedirs(LOG_DIR, exist_ok=True)

# 로그 파일 설정 (재기동 시 초기화)
log_file = os.path.join(LOG_DIR, "app.log")

# 파일 핸들러 (재기동 시 초기화를 위해 mode='w' 사용)
file_handler = logging.FileHandler(log_file, mode='w', encoding="utf-8")
file_handler.setLevel(logging.INFO)
file_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
file_handler.setFormatter(file_formatter)

# 콘솔 핸들러
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
console_handler.setFormatter(console_formatter)

# 루트 로거 설정
root_logger = logging.getLogger()
root_logger.setLevel(logging.INFO)
root_logger.addHandler(file_handler)
root_logger.addHandler(console_handler)

logger = logging.getLogger(__name__)
logger.info(f"📝 로그 파일 설정 완료: {log_file}")

# 전역 동기화 상태 변수
sync_status = {
    "is_running": False,
    "progress": 0,
    "total_epics": 0,
    "completed_epics": 0,
    "failed_epics": 0,
    "current_epic": "",
    "message": "",
    "failed_list": []
}

# Create FastAPI app
app = FastAPI()

# 스케줄러 초기화 (SSL 문제로 임시 비활성화)
# scheduler = BackgroundScheduler()

# 기존 관리자 페이지 제거됨

@app.get("/")
async def root():
    """메인 페이지 - effort-management로 리다이렉트"""
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/effort-management/effort-management.html")

@app.get("/effort")
async def effort_management():
    """기존 /effort 경로 - effort-management로 리다이렉트"""
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/effort-management/effort-management.html")

@app.get("/static/effort-management.html")
async def static_effort_redirect():
    """기존 /static 경로 - effort-management로 리다이렉트"""
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/effort-management/effort-management.html")

@app.on_event("startup")
async def startup_event():
    try:
        logger.info("=" * 80)
        logger.info("🚀 서버 시작 중...")
        logger.info("=" * 80)
        
        # 벡터 DB 색인 스킵 (동기화 시 증분 색인으로 처리)
        logger.info("📚 [1/3] 벡터 DB 색인 확인...")
        try:
            json_file_path = os.path.join(DOCS_DIR, "effort_estimations.json")
            if os.path.exists(json_file_path):
                # JSON 파일 정보 확인
                import json
                with open(json_file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    total_items = len(data)
                
                logger.info(f"   📄 effort_estimations.json 파일 확인: {total_items}개 항목")
                logger.info(f"   ℹ️ 색인은 Epic 동기화 시 자동으로 실행됩니다 (증분 색인)")
                logger.info(f"   ℹ️ 수동 재색인이 필요하면 웹 UI에서 '데이터 재색인' 버튼 클릭")
            else:
                logger.warning("   ⚠️ effort_estimations.json 파일 없음")
        except Exception as check_error:
            logger.error(f"   ❌ 파일 확인 실패: {str(check_error)}")
        
        # 카테고리 자동 마이그레이션
        logger.info("📂 [2/3] 카테고리 자동 마이그레이션 중...")
        await auto_migrate_categories()
        logger.info("   ✅ 카테고리 마이그레이션 완료")
        
        # Epic 자동 동기화 스케줄러 시작 (SSL 문제로 임시 비활성화)
        logger.info("⏰ [3/3] 스케줄러 설정 중...")
        # scheduler.add_job(
        #     sync_completed_epics_background,
        #     'cron',
        #     hour=3,
        #     minute=0,
        #     id='auto_sync_completed_epics',
        #     replace_existing=True
        # )
        # scheduler.start()
        logger.info("   ⚠️ Epic 자동 동기화 스케줄러 비활성화 (SSL 문제)")
        logger.info("   ℹ️ 수동 실행만 가능합니다")
        
        logger.info("=" * 80)
        logger.info("✅ 서버 기동 완료! 🎉")
        logger.info("=" * 80)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        logger.error(f"❌ Error during startup: {str(e)}")

@app.on_event("shutdown")
async def shutdown_event():
    """서버 종료 시 스케줄러 정리"""
    try:
        # scheduler.shutdown()  # SSL 문제로 임시 비활성화
        logger.info("✅ 서버 종료 완료")
    except Exception as e:
        logger.error(f"❌ 서버 종료 오류: {str(e)}")

async def auto_migrate_categories():
    """카테고리 변경 시 자동 마이그레이션"""
    try:
        logger.info("📊 카테고리 변경 감지 중...")
        
        # 현재 카테고리 로드
        from ..services.effort_estimation import CategoryManager
        category_manager = CategoryManager()
        current_categories = category_manager.get_categories()
        
        # 기존 카테고리 파일 로드 (categories.json)
        import os
        categories_file = os.path.join(DOCS_DIR, "categories.json")
        
        if not os.path.exists(categories_file):
            logger.warning("카테고리 파일을 찾을 수 없습니다")
            return
        
        # 모든 공수 산정 데이터 가져오기
        estimations = effort_manager.get_all_estimations()
        
        # 카테고리 변경 감지 및 자동 마이그레이션
        updated_count = 0
        reset_count = 0
        
        for estimation in estimations:
            if not estimation.major_category or not estimation.minor_category or not estimation.sub_category:
                continue
            
            # 현재 카테고리 구조에서 해당 카테고리가 존재하는지 확인
            major = estimation.major_category
            minor = estimation.minor_category
            sub = estimation.sub_category
            
            # 대중소분류가 모두 정확히 일치하는지 확인
            is_valid = (
                major in current_categories and
                minor in current_categories.get(major, {}) and
                sub in current_categories.get(major, {}).get(minor, [])
            )
            
            if is_valid:
                # 정확히 일치하는 경우: 그대로 유지 (변경 없음)
                logger.debug(f"✅ 카테고리 유지: {major} > {minor} > {sub}")
            else:
                # 하나라도 안 맞는 경우: 카테고리 초기화 (사용자가 다시 선택하도록)
                logger.info(f"🔄 카테고리 초기화: {major} > {minor} > {sub} -> (초기화됨, 재선택 필요)")
                estimation.major_category = None
                estimation.minor_category = None
                estimation.sub_category = None
                reset_count += 1
        
        if reset_count > 0:
            # 변경사항 저장
            effort_manager.save_data()
            logger.info(f"✅ 카테고리 자동 마이그레이션 완료: {reset_count}개 데이터 카테고리 초기화 (재선택 필요)")
        else:
            logger.info("📊 카테고리 변경 사항 없음 (모든 카테고리가 정확히 일치함)")
        
    except Exception as e:
        logger.error(f"❌ 카테고리 자동 마이그레이션 오류: {str(e)}")

@app.post("/upload_pdf/")
async def upload_pdf(file: UploadFile = File(...)):
    try:
        file_path = os.path.join(DOCS_DIR, file.filename)
        with open(file_path, "wb") as f:
            shutil.copyfileobj(file.file, f)
            
        if index_document(file_path, "pdf"):
            return {"message": f"'{file.filename}' indexed successfully"}
        else:
            return JSONResponse(status_code=500, content={"error": "Failed to index document"})
            
    except Exception as e:
        logger.error(f"❌ Error uploading PDF: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

def extract_questions(text: str) -> set:
    """Q1:, Q2: 형식으로 시작하는 질문 추출"""
    return set(
        m.strip()
        for m in re.findall(r"Q\d+:\s*(.+?)(?:\n|$)", text)
    )


@app.post("/upload_text/")
async def upload_text(text: str = Form(...), source: str = Form(...)):
    try:
        # 1. 입력값 검증
        if not text.strip():
            return JSONResponse(status_code=400, content={"error": "텍스트 내용이 비어있습니다."})

        safe_source = re.sub(r'[\\/]', '_', source.strip()) or "TEMP"
        txt_filename = f"{safe_source}.txt"
        txt_path = os.path.join(DOCS_DIR, txt_filename)

        logger.info(f"📝 텍스트 업로드 요청 - source: {safe_source}")

        # 2. 중복 질문 필터링
        new_questions = extract_questions(text.strip())
        existing_text = ""
        existing_questions = set()

        if os.path.exists(txt_path):
            with open(txt_path, "r", encoding="utf-8") as f:
                existing_text = f.read()
                existing_questions = extract_questions(existing_text)

        duplicated = new_questions & existing_questions
        if duplicated:
            return JSONResponse(
                status_code=200,
                content={
                    "message": "일부 또는 전체 질문이 이미 존재합니다.",
                    "duplicated_questions": list(duplicated)
                }
            )

        # 3. 텍스트 파일에 append 저장
        with open(txt_path, "a", encoding="utf-8") as f:
            if existing_text:
                f.write("\n")
            f.write(text.strip())

        # 4. 색인 처리 (database.py의 index_document 호출)
        if index_document(txt_path, file_type="txt", force=True):
            logger.info(f"✅ '{safe_source}' 텍스트 색인 완료")
            return {
                "message": f"'{safe_source}' 텍스트가 성공적으로 추가되고 재색인되었습니다.",
            }
        else:
            return JSONResponse(status_code=500, content={"error": "문서 색인 실패"})

    except Exception as e:
        logger.error(f"❌ upload_text 오류: {str(e)}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/indexed_files/")
async def get_indexed_files_endpoint():
    try:
        files = get_indexed_files()
        # Add download URLs for each file
        files_with_urls = [
            {
                "filename": filename,
                "download_url": f"/download/{filename}"
            }
            for filename in files
        ]
        return {"indexed_files": files_with_urls}
    except Exception as e:
        logger.error(f"❌ Error getting indexed files: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/download/{filename}")
async def download_file(filename: str):
    try:
        # Validate file extension
        if not filename.endswith((".pdf", ".txt")):
            return JSONResponse(
                status_code=400,
                content={"error": "Only .pdf and .txt files are supported"}
            )
            
        file_path = os.path.join(DOCS_DIR, filename)
        
        if not os.path.exists(file_path):
            return JSONResponse(
                status_code=404,
                content={"error": f"File '{filename}' not found"}
            )
            
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type="application/octet-stream"
        )
            
    except Exception as e:
        logger.error(f"❌ Error downloading file: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )

@app.post("/ask/")
async def ask_question(question: str = Form(...)):
    try:
        logger.info(f"💬 Question received: {question}")
        
        # 일반 질문은 모의 응답 사용
        result = mock_qa_response(question)

        if "error" in result:
            return JSONResponse(status_code=400, content={"error": result["error"]})
        
        sources_text = format_sources(result["sources"])

        return {
            "question": result["question"],
            "answer": result["answer"],
            # "sources": result["sources"],
            "formatted_response": f"{result['answer']}{sources_text}"
        }

    except Exception as e:
        logger.error(f"❌ Error processing question: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/slack/test")
async def slack_test():
    """슬랙봇 연결 테스트"""
    return {"status": "success", "message": "슬랙봇이 정상적으로 연결되었습니다!"}

@app.post("/slack/test")
async def slack_test_post(request: Request):
    """슬랙 URL 검증 테스트"""
    try:
        data = await request.json()
        logger.info(f"🧪 테스트 POST 요청 수신: {data}")
        
        if data.get("type") == "url_verification":
            challenge = data.get("challenge")
            logger.info(f"🔐 테스트 URL 인증 - Challenge: {challenge}")
            return {"challenge": challenge}
        
        return {"status": "success", "received": data}
    except Exception as e:
        logger.error(f"❌ 테스트 POST 오류: {str(e)}")
        return {"status": "error", "message": str(e)}

@app.post("/slack/events")
async def slack_event_listener(
    request: Request,
    background_tasks: BackgroundTasks,
    x_slack_retry_num: str = Header(default=None),
    x_slack_retry_reason: str = Header(default=None)
):
    try:
        # 요청 로깅
        logger.info(f"🔔 Slack 이벤트 수신 - Retry: {x_slack_retry_num}, Reason: {x_slack_retry_reason}")
        
        data = await request.json()
        logger.info(f"📦 수신된 데이터: {data}")

        # 중복 전송 방지
        if x_slack_retry_num:
            logger.info("⏭️ 중복 요청으로 인한 스킵")
            return {"status": "ok"}

        # Slack URL 인증
        if data.get("type") == "url_verification":
            challenge = data.get("challenge")
            logger.info(f"🔐 URL 인증 요청 - Challenge: {challenge}")
            return {"challenge": challenge}

        # 실제 이벤트 처리
        if data.get("type") == "event_callback":
            event = data.get("event", {})
            event_type = event.get("type")
            user = event.get("user", "알 수 없음")
            
            logger.info(f"📨 이벤트 타입: {event_type}, 사용자: {user}")

            # 앱 멘션
            if event_type == "app_mention":
                channel = event.get("channel")
                thread_ts = event.get("thread_ts", event.get("ts"))
                text = clean_mention(event.get("text", ""))
                logger.info(f"📥 채널 수신된 메시지 {user}: {text}")
                background_tasks.add_task(handle_slack_message, text, channel, thread_ts, event.get("ts"))

            # DM 메시지
            elif event_type == "message" and event.get("channel_type") == "im" and not event.get("bot_id"):
                channel = event.get("channel")
                text = clean_mention(event.get("text", ""))  # 멘션 제거 추가
                logger.info(f"📥 앱 메세지 탭 수신된 메시지 {user} : {text}")
                background_tasks.add_task(handle_slack_message, text, channel, None, event.get("ts"))
            
            # 이모지 리액션 이벤트 (피드백 수집)
            elif event_type == "reaction_added":
                logger.info(f"👍 reaction_added 이벤트 수신! reaction={event.get('reaction')}, item={event.get('item')}")
                from ..utils.slack import handle_slack_reaction
                background_tasks.add_task(handle_slack_reaction, event)
            
            else:
                logger.info(f"ℹ️ 처리되지 않은 이벤트 타입: {event_type}")

        logger.info("✅ Slack 이벤트 처리 완료")
        return {"status": "ok"}

    except Exception as e:
        logger.error(f"❌ Error handling Slack event: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})





@app.delete("/files/{filename}")
async def delete_file(filename: str):
    try:
        # Validate file extension
        if not filename.endswith((".pdf", ".txt")):
            return JSONResponse(
                status_code=400,
                content={"error": "Only .pdf and .txt files are supported"}
            )
            
        file_path = os.path.join(DOCS_DIR, filename)
        
        if remove_document(file_path):
            return {"message": f"File '{filename}' deleted successfully"}
        else:
            return JSONResponse(
                status_code=404,
                content={"error": f"File '{filename}' not found"}
            )
            
    except Exception as e:
        logger.error(f"❌ Error deleting file: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={"error": str(e)}
        )

@app.post("/indexed_files")
async def reindex_all_files():
    try:
        logger.info("🔄 Starting complete reindexing process...")
        
        # First, reset the Chroma DB
        if not reset_vectordb():
            return JSONResponse(
                status_code=500,
                content={"error": "Failed to reset database"}
            )
        
        logger.info("🗑️ Database reset complete, starting reindexing...")
        indexed_count = 0
        error_count = 0
        
        # Get list of all files
        for filename in os.listdir(DOCS_DIR):
            if filename.endswith((".pdf", ".txt")):
                file_path = os.path.join(DOCS_DIR, filename)
                file_type = "pdf" if filename.endswith(".pdf") else "txt"
                
                try:
                    # No need for force=True since DB is fresh
                    if index_document(file_path, file_type):
                        indexed_count += 1
                        logger.info(f"✅ Indexed: {filename}")
                    else:
                        error_count += 1
                        logger.error(f"❌ Failed to index: {filename}")
                except Exception as e:
                    error_count += 1
                    logger.error(f"❌ Error indexing {filename}: {str(e)}")
        
        message = f"전체 재색인 완료: {indexed_count}개 성공"
        if error_count > 0:
            message += f", {error_count}개 실패"
            
        logger.info(message)
        return {"message": message}
        
    except Exception as e:
        error_msg = f"❌ Error during reindexing: {str(e)}"
        logger.error(error_msg)
        return JSONResponse(status_code=500, content={"error": error_msg})

@app.post("/index_url/")
async def upload_url(url: str = Form(...), source: str = Form(default="web")):
    try:
        logger.info(f"🌐 URL 크롤링 요청: {url}")

        # ✅ 웹 페이지 요청 및 파싱
        response = requests.get(url, timeout=10)
        if response.status_code != 200:
            return JSONResponse(status_code=400, content={"error": f"Failed to fetch URL: {response.status_code}"})

        soup = BeautifulSoup(response.text, "html.parser")

        # ✅ 주요 태그 위주로 텍스트 구조화
        lines = []

        # 제목 계열 먼저
        for header in soup.find_all(['h1', 'h2', 'h3', 'h4']):
            lines.append(f"# {header.get_text(strip=True)}")

        # 단락
        for paragraph in soup.find_all('p'):
            lines.append(paragraph.get_text(strip=True))

        # 리스트
        for li in soup.find_all('li'):
            lines.append(f"- {li.get_text(strip=True)}")

        # 기타 텍스트 누락 방지용 (기본적 body에서 추가로 가져오기)
        body_text = soup.body.get_text(separator="\n", strip=True) if soup.body else ""
        lines.append(body_text)

        # 중복 제거 및 정리
        clean_lines = []
        for line in lines:
            line = line.strip()
            if line and line not in clean_lines:
                clean_lines.append(line)

        text = "\n".join(clean_lines)

        # ✅ 파일 저장
        safe_source = re.sub(r'[\\/]', '_', source.strip()) or "web"
        txt_filename = f"{safe_source}.txt"
        txt_path = os.path.join(DOCS_DIR, txt_filename)
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(text)

        # ✅ 색인 처리
        if index_document(txt_path, "txt", force=True):
            return {"message": f"'{url}' 크롤링 및 색인 성공", "source": txt_filename}
        else:
            return JSONResponse(status_code=500, content={"error": "문서 색인 실패"})

    except Exception as e:
        logger.error(f"❌ upload_url 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

# ask_preview 엔드포인트 제거됨

# ==================== 공수 산정 관련 엔드포인트 ====================

@app.post("/effort/add/")
async def add_effort_estimation(
    jira_ticket: str = Form(...),
    title: str = Form(...),
    story_points: float = Form(...),
    estimation_reason: str = Form(default=None),
    tech_stack: str = Form(default=None),
    team_member: str = Form(default=None),
    notes: str = Form(default=None),
    major_category: str = Form(default=None),
    minor_category: str = Form(default=None),
    sub_category: str = Form(default=None),
    auto_classify: bool = Form(default=False)
):
    """수동으로 공수 산정 데이터 추가 (Story Point 기반)"""
    try:
        # 기술 스택 파싱
        tech_stack_list = None
        if tech_stack:
            tech_stack_list = [tech.strip() for tech in tech_stack.split(',')]
        
        # 자동 분류 활성화 시
        if auto_classify:
            predicted_category, confidence = auto_classify(title)
            if predicted_category and confidence > 0.5:
                logger.info(f"자동 분류 결과: {predicted_category} (신뢰도: {confidence:.2f})")
                # 예측된 카테고리를 사용
                category_parts = predicted_category.split(' > ')
                if len(category_parts) >= 3:
                    major_category = category_parts[0]
                    minor_category = category_parts[1]
                    sub_category = category_parts[2]
            else:
                logger.warning(f"자동 분류 신뢰도 부족: {confidence:.2f}")
        
        estimation = EffortEstimation(
            jira_ticket=jira_ticket,
            title=title,
            story_points=story_points,
            estimation_reason=estimation_reason,
            tech_stack=tech_stack_list,
            team_member=team_member,
            notes=notes,
            major_category=major_category,
            minor_category=minor_category,
            sub_category=sub_category
        )
        
        if effort_manager.add_estimation(estimation):
            # 공수 산정 데이터를 색인에 추가
            effort_text = effort_manager.format_for_indexing()
            effort_file_path = os.path.join(DOCS_DIR, "effort_estimations.txt")
            with open(effort_file_path, "w", encoding="utf-8") as f:
                f.write(effort_text)
            
            # 색인 업데이트
            index_document(effort_file_path, "txt", force=True)
            
            return {"message": f"공수 산정 데이터가 성공적으로 추가되었습니다: {title}"}
        else:
            return JSONResponse(status_code=500, content={"error": "공수 산정 데이터 추가 실패"})
            
    except Exception as e:
        logger.error(f"❌ 공수 산정 데이터 추가 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

def save_web_qa_mapping(question: str, answer: str, sources: list = None):
    """웹 QA 매핑 저장"""
    try:
        web_mapping_file = os.path.join(DOCS_DIR, "web_qa_mapping.json")
        
        # 기존 매핑 로드
        web_qa_mapping = {}
        if os.path.exists(web_mapping_file):
            try:
                with open(web_mapping_file, 'r', encoding='utf-8') as f:
                    web_qa_mapping = json.load(f)
            except (json.JSONDecodeError, Exception) as e:
                logger.warning(f"⚠️ 웹 QA 매핑 파일 읽기 오류: {e}, 빈 딕셔너리로 시작")
                web_qa_mapping = {}
        
        # 새로운 QA 항목 추가 (타임스탬프를 키로 사용)
        qa_id = datetime.now().isoformat()
        web_qa_mapping[qa_id] = {
            "question": question,
            "answer": answer,
            "sources": sources or [],
            "timestamp": qa_id,
            "source": "web"
        }
        
        # 파일 저장
        with open(web_mapping_file, 'w', encoding='utf-8') as f:
            json.dump(web_qa_mapping, f, ensure_ascii=False, indent=2)
        
        logger.info(f"💾 웹 QA 매핑 저장: {question[:30]}...")
        return True
    except Exception as e:
        logger.error(f"❌ 웹 QA 매핑 저장 오류: {str(e)}")
        return False

@app.post("/effort/ask/")
async def ask_effort_question(question: str = Form(...)):
    """공수 산정 관련 질문"""
    try:
        logger.info(f"💬 공수 산정 질문 수신: {question}")
        
        try:
            result = run_effort_qa_chain(question)
        except Exception as e:
            if "quota" in str(e).lower() or "insufficient_quota" in str(e).lower():
                logger.warning("⚠️ OpenAI API 할당량 초과, 공수 산정 모의 응답 사용")
                result = mock_effort_qa_response(question)
            else:
                raise e
        
        if "error" in result:
            return JSONResponse(status_code=400, content={"error": result["error"]})
        
        # 웹 QA 매핑 저장
        save_web_qa_mapping(
            question=result["question"],
            answer=result["answer"],
            sources=result.get("sources", [])
        )
        
        sources_text = format_sources(result["sources"])
        
        return {
            "question": result["question"],
            "answer": result["answer"],
            "formatted_response": f"{result['answer']}{sources_text}",
            "feedback_enabled": result.get("feedback_enabled", False),
            "search_session_id": result.get("search_session_id", ""),
            "sources": result.get("sources", [])
        }
        
    except Exception as e:
        logger.error(f"❌ 공수 산정 질문 처리 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/effort/ask-feedback/")
async def ask_effort_question_with_feedback(request: dict):
    """피드백 기반 공수 산정 질문 재검색"""
    try:
        question = request.get("question", "")
        excluded_sources = request.get("excluded_sources", [])
        
        logger.info(f"🔄 피드백 기반 공수 산정 질문 수신: {question}")
        logger.info(f"🚫 제외할 소스: {excluded_sources}")
        
        try:
            result = run_effort_qa_with_feedback(question, excluded_sources)
        except Exception as e:
            if "quota" in str(e).lower() or "insufficient_quota" in str(e).lower():
                logger.warning("⚠️ OpenAI API 할당량 초과, 공수 산정 모의 응답 사용")
                result = mock_effort_qa_response(question)
            else:
                raise e
        
        if "error" in result:
            return JSONResponse(status_code=400, content={"error": result["error"]})
        
        sources_text = format_sources(result["sources"])
        
        return {
            "question": result["question"],
            "answer": result["answer"],
            "formatted_response": f"{result['answer']}{sources_text}",
            "feedback_enabled": result.get("feedback_enabled", False),
            "search_session_id": result.get("search_session_id", ""),
            "sources": result.get("sources", []),
            "is_feedback_search": result.get("is_feedback_search", False)
        }
        
    except Exception as e:
        logger.error(f"❌ 피드백 기반 공수 산정 질문 처리 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/effort/statistics/")
async def get_effort_statistics_endpoint():
    """공수 산정 통계 조회"""
    try:
        stats = get_effort_statistics()
        return stats
    except Exception as e:
        logger.error(f"❌ 공수 산정 통계 조회 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/effort/feedback-statistics/weekly-positive-ratio/")
async def get_feedback_weekly_positive_ratio_endpoint():
    """주 단위 긍정 피드백 비율 통계 조회"""
    try:
        from ..services.effort_qa import get_feedback_weekly_positive_ratio
        stats = get_feedback_weekly_positive_ratio()
        return stats
    except Exception as e:
        logger.error(f"❌ 주 단위 피드백 통계 조회 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/effort/search/{feature_name}")
async def search_effort_features(feature_name: str):
    """기능명으로 공수 산정 데이터 검색"""
    try:
        results = search_similar_features(feature_name)
        return {"feature_name": feature_name, "results": results}
    except Exception as e:
        logger.error(f"❌ 공수 산정 검색 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/effort/debug-search/")
async def debug_search_effort_features(query: str):
    """디버깅용 공수 산정 데이터 검색"""
    try:
        logger.info(f"🔍 디버깅 검색 요청: '{query}'")
        
        # 직접 데이터 검색
        estimations = effort_manager.get_all_estimations()
        logger.info(f"📊 전체 데이터 수: {len(estimations)}")
        
        # 제목으로 검색
        matching_estimations = []
        for est in estimations:
            if query.lower() in est.title.lower():
                matching_estimations.append({
                    "jira_ticket": est.jira_ticket,
                    "title": est.title,
                    "story_points": est.story_points
                })
        
        logger.info(f"🔍 매칭된 데이터 수: {len(matching_estimations)}")
        
        return {
            "query": query,
            "total_estimations": len(estimations),
            "matching_estimations": matching_estimations,
            "all_titles": [est.title for est in estimations[:10]]  # 처음 10개 제목만
        }
    except Exception as e:
        logger.error(f"❌ 디버깅 검색 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/effort/vector-status/")
async def get_vector_db_status():
    """벡터 DB 상태 확인"""
    try:
        logger.info("🔍 벡터 DB 상태 확인 시작")
        
        vectordb = get_vectordb()
        collection = vectordb.get()
        
        # 소스별 문서 수 집계
        source_counts = {}
        for metadata in collection["metadatas"]:
            if isinstance(metadata, dict):
                source = metadata.get("source", "unknown")
                source_counts[source] = source_counts.get(source, 0) + 1
        
        logger.info(f"📊 벡터 DB 전체 문서 수: {len(collection['ids'])}")
        logger.info(f"📊 소스별 문서 수: {source_counts}")
        
        return {
            "total_documents": len(collection['ids']),
            "source_counts": source_counts,
            "sources": list(source_counts.keys())
        }
    except Exception as e:
        logger.error(f"❌ 벡터 DB 상태 확인 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/effort/cleanup-temp/")
async def cleanup_temp_files():
    """TEMP.txt 파일 벡터 DB에서 제거"""
    try:
        logger.info("🧹 TEMP.txt 파일 정리 시작")
        
        vectordb = get_vectordb()
        collection = vectordb.get()
        
        # TEMP.txt 관련 문서 ID 찾기
        temp_doc_ids = []
        for i, metadata in enumerate(collection["metadatas"]):
            if isinstance(metadata, dict) and metadata.get("source") == "TEMP.txt":
                temp_doc_ids.append(collection["ids"][i])
        
        if temp_doc_ids:
            # TEMP.txt 문서들 삭제
            vectordb._collection.delete(temp_doc_ids)
            logger.info(f"🗑️ TEMP.txt 문서 {len(temp_doc_ids)}개 삭제 완료")
            
            return {
                "message": f"TEMP.txt 파일 {len(temp_doc_ids)}개 문서가 벡터 DB에서 제거되었습니다.",
                "removed_count": len(temp_doc_ids)
            }
        else:
            return {
                "message": "TEMP.txt 파일이 벡터 DB에 존재하지 않습니다.",
                "removed_count": 0
            }
            
    except Exception as e:
        logger.error(f"❌ TEMP.txt 정리 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/effort/reindex/")
async def reindex_effort_data():
    """공수 산정 데이터 재인덱싱"""
    try:
        logger.info("🔄 공수 산정 데이터 재인덱싱 시작")
        
        # effort_estimations.txt 파일 재인덱싱
        effort_file_path = os.path.join(DOCS_DIR, "effort_estimations.txt")
        
        if os.path.exists(effort_file_path):
            logger.info(f"📄 effort_estimations.txt 파일 발견: {effort_file_path}")
            
            # 강제 재인덱싱
            if index_document(effort_file_path, file_type="txt", force=True):
                logger.info("✅ effort_estimations.txt 재인덱싱 완료")
                
                # 벡터 DB 상태 확인
                vectordb = get_vectordb()
                collection = vectordb.get()
                logger.info(f"📊 벡터 DB 문서 수: {len(collection['ids'])}")
                
                # effort_estimations.txt 관련 문서 수 확인
                effort_docs = 0
                for metadata in collection["metadatas"]:
                    if isinstance(metadata, dict) and metadata.get("source") == "effort_estimations.txt":
                        effort_docs += 1
                
                logger.info(f"📊 effort_estimations.txt 문서 수: {effort_docs}")
                
                return {
                    "message": "공수 산정 데이터 재인덱싱 완료",
                    "total_documents": len(collection['ids']),
                    "effort_documents": effort_docs
                }
            else:
                return JSONResponse(status_code=500, content={"error": "재인덱싱 실패"})
        else:
            return JSONResponse(status_code=404, content={"error": "effort_estimations.txt 파일을 찾을 수 없습니다"})
            
    except Exception as e:
        logger.error(f"❌ 재인덱싱 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/effort/sync-jira/")
async def sync_jira_data(request: Request):
    """Jira 티켓 데이터 동기화"""
    try:
        # 요청 데이터 로깅
        logger.info(f"🔄 Jira 동기화 요청 수신 시작")
        
        # Content-Type 확인
        content_type = request.headers.get("content-type", "")
        logger.info(f"🔄 Content-Type: {content_type}")
        
        # FormData 파싱
        form_data = await request.form()
        logger.info(f"🔄 FormData keys: {list(form_data.keys())}")
        
        ticket_key = form_data.get("ticket_key")
        major_category = form_data.get("major_category")
        minor_category = form_data.get("minor_category")
        sub_category = form_data.get("sub_category")
        logger.info(f"🔄 Jira 동기화 요청 수신: ticket_key={ticket_key}, categories={major_category}/{minor_category}/{sub_category}")
        
        if not ticket_key:
            logger.error("❌ ticket_key 파라미터가 없습니다")
            return JSONResponse(status_code=422, content={"error": "ticket_key 파라미터가 필요합니다"})
        
        jira = create_jira_integration()
        if not jira:
            logger.error("❌ Jira 설정이 없습니다")
            return JSONResponse(status_code=400, content={"error": "Jira 설정이 필요합니다"})
        
        if not jira.test_connection():
            logger.error("❌ Jira 연결 실패")
            return JSONResponse(status_code=400, content={"error": "Jira 연결에 실패했습니다"})
        
        logger.info(f"🔄 티켓 '{ticket_key}' 동기화 시작")
        result = jira.sync_ticket_data(ticket_key, major_category, minor_category, sub_category)
        
        if result["success"]:
            # 색인 업데이트
            effort_text = effort_manager.format_for_indexing()
            effort_file_path = os.path.join(DOCS_DIR, "effort_estimations.txt")
            with open(effort_file_path, "w", encoding="utf-8") as f:
                f.write(effort_text)
            
            index_document(effort_file_path, "txt", force=True)
            
            logger.info(f"✅ 티켓 '{ticket_key}' 동기화 완료")
            return {"message": f"티켓 '{ticket_key}' 데이터 동기화 완료"}
        else:
            # 티켓 타입 필터링인 경우 특별한 메시지 반환
            if result["reason"] == "not_found_or_invalid_type":
                logger.warning(f"⚠️ 허용되지 않은 티켓 타입: {ticket_key}")
                return JSONResponse(
                    status_code=400, 
                    content={"error": f"Epic 타입의 티켓은 동기화할 수 없습니다. 허용된 타입: 작업, 스토리, 버그, Story, Task, Bug"}
                )
            elif result["reason"] == "no_estimation_data":
                logger.warning(f"⚠️ 공수 데이터 없음: {ticket_key}")
                return JSONResponse(status_code=400, content={"error": f"티켓 '{ticket_key}'에서 공수 데이터를 추출할 수 없습니다"})
            else:
                logger.error(f"❌ 티켓 '{ticket_key}' 동기화 실패")
                return JSONResponse(status_code=500, content={"error": "Jira 데이터 동기화 실패"})
            
    except Exception as e:
        logger.error(f"❌ Jira 동기화 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/effort/list/")
async def list_effort_estimations(
    major_category: str = None,
    minor_category: str = None,
    sub_category: str = None,
    search: str = None,
    page: int = 1,
    page_size: int = 100
):
    """공수 산정 데이터 목록 조회 (카테고리 필터링 지원)"""
    try:
        estimations = effort_manager.get_all_estimations()
        
        # 검색 적용 (제목 또는 Jira 티켓)
        if search:
            search_term = search.lower().strip()
            filtered_estimations = []
            for estimation in estimations:
                # 제목에서 검색
                if search_term in estimation.title.lower():
                    filtered_estimations.append(estimation)
                # Jira 티켓에서 검색
                elif estimation.jira_ticket and search_term in estimation.jira_ticket.lower():
                    filtered_estimations.append(estimation)
            estimations = filtered_estimations
        
        # 페이징 처리
        total_count = len(estimations)
        total_pages = (total_count + page_size - 1) // page_size  # 올림 계산
        
        # 페이지 범위 계산
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        
        # 현재 페이지 데이터 추출
        paginated_estimations = estimations[start_index:end_index]
        
        # 넘버링 추가 (전체 데이터 기준)
        for i, estimation in enumerate(paginated_estimations):
            estimation.sequence_number = start_index + i + 1
        
        # 데이터 목록용: description과 comments 제외 (응답 크기 축소)
        estimations_data = []
        for estimation in paginated_estimations:
            est_dict = estimation.__dict__.copy()
            # description과 comments 제외 (긴 텍스트)
            est_dict.pop('description', None)
            est_dict.pop('comments', None)
            estimations_data.append(est_dict)
        
        jira_url = os.getenv('JIRA_URL', 'https://enomix.atlassian.net')
        return {
            "estimations": estimations_data,
            "jira_url": jira_url,
            "pagination": {
                "current_page": page,
                "page_size": page_size,
                "total_count": total_count,
                "total_pages": total_pages,
                "has_previous": page > 1,
                "has_next": page < total_pages
            }
        }
    except Exception as e:
        logger.error(f"❌ 공수 산정 목록 조회 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

# 카테고리 관리 API
@app.post("/effort/auto-classify/")
async def auto_classify_estimations():
    """미분류 데이터 자동 분류"""
    try:
        estimations = effort_manager.get_all_estimations()
        
        # 미분류 데이터 필터링 (카테고리가 없는 경우)
        unclassified = [
            est for est in estimations 
            if not est.major_category or not est.minor_category or not est.sub_category
        ]
        
        logger.info(f"미분류 데이터: {len(unclassified)}개")
        
        # 자동 분류 실행
        classified_count = 0
        total_confidence = 0
        
        # 신뢰도별 카테고리 분류
        low_confidence = []  # 0.1 ~ 0.3
        medium_confidence = []  # 0.3 ~ 0.5
        high_confidence = []  # 0.5 이상
        
        for estimation in unclassified:
            # 제목과 설명을 모두 사용하여 분류
            classification_text = estimation.title
            if estimation.notes:
                classification_text += " " + estimation.notes
            
            predicted_category, confidence = auto_classify(classification_text)
            
            # confidence를 명시적으로 float로 변환
            try:
                conf_float = float(confidence) if confidence is not None else 0.0
                conf_str = f"{conf_float:.2f}"
            except Exception as e:
                logger.error(f"❌ confidence 변환 오류: {e}, confidence={confidence}, type={type(confidence)}")
                conf_float = 0.0
                conf_str = "0.00"
            
            if predicted_category and conf_float >= 0.5:
                # 높은 신뢰도: 자동 적용
                category_parts = predicted_category.split(' > ')
                if len(category_parts) >= 3:
                    estimation.major_category = category_parts[0]
                    estimation.minor_category = category_parts[1]
                    estimation.sub_category = category_parts[2]
                    classified_count += 1
                    total_confidence += conf_float
                    high_confidence.append((estimation.title, predicted_category, conf_float))
                    logger.info(f"✅ 자동 분류 (높음): {estimation.title} -> {predicted_category} (신뢰도: {conf_str})")
            elif predicted_category and conf_float >= 0.3:
                # 중간 신뢰도: 사용자 확인 후 적용
                medium_confidence.append((estimation.title, predicted_category, conf_float))
                logger.info(f"⚠️ 신뢰도 중간: {estimation.title} -> {predicted_category} (신뢰도: {conf_str})")
            elif predicted_category and conf_float >= 0.1:
                # 낮은 신뢰도: 제안만
                low_confidence.append((estimation.title, predicted_category, conf_float))
                logger.info(f"📝 신뢰도 낮음: {estimation.title} -> {predicted_category} (신뢰도: {conf_str})")
            else:
                logger.info(f"❌ 분류 실패: {estimation.title} (신뢰도: {conf_str if conf_float else 'N/A'})")
        
        # 평균 신뢰도 계산
        avg_confidence = total_confidence / classified_count if classified_count > 0 else 0
        
        # 변경사항 저장
        effort_manager.save_data()
        
        # 튜플을 딕셔너리로 변환 (JSON 직렬화 가능하도록)
        def tuple_to_dict(tup_list):
            return [
                {
                    "title": str(tup[0]),
                    "category": str(tup[1]),
                    "confidence": round(float(tup[2]), 2)
                }
                for tup in tup_list
            ]
        
        return {
            "message": "자동 분류 완료",
            "total_unclassified": len(unclassified),
            "high_confidence_count": len(high_confidence),
            "medium_confidence_count": len(medium_confidence),
            "low_confidence_count": len(low_confidence),
            "classified_count": classified_count,
            "average_confidence": round(avg_confidence, 2),
            "high_confidence": tuple_to_dict(high_confidence[:5]),
            "medium_confidence": tuple_to_dict(medium_confidence[:5]),
            "low_confidence": tuple_to_dict(low_confidence[:5])
        }
        
    except Exception as e:
        logger.error(f"❌ 자동 분류 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/effort/categories/")
async def get_categories():
    """카테고리 구조 조회"""
    try:
        from ..services.effort_estimation import CategoryManager
        category_manager = CategoryManager()
        return category_manager.get_categories()
    except Exception as e:
        logger.error(f"❌ 카테고리 조회 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/effort/categories/major/")
async def get_major_categories():
    """대분류 목록 조회"""
    try:
        from ..services.effort_estimation import CategoryManager
        category_manager = CategoryManager()
        return {"categories": category_manager.get_major_categories()}
    except Exception as e:
        logger.error(f"❌ 대분류 조회 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/effort/categories/minor/")
async def get_minor_categories(major: str):
    """중분류 목록 조회"""
    try:
        from ..services.effort_estimation import CategoryManager
        category_manager = CategoryManager()
        return {"categories": category_manager.get_minor_categories(major)}
    except Exception as e:
        logger.error(f"❌ 중분류 조회 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/effort/categories/sub/")
async def get_sub_categories(major: str, minor: str):
    """소분류 목록 조회"""
    try:
        from ..services.effort_estimation import CategoryManager
        category_manager = CategoryManager()
        return {"categories": category_manager.get_sub_categories(major, minor)}
    except Exception as e:
        logger.error(f"❌ 소분류 조회 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/effort/categories/")
async def add_category(request: Request):
    """새 카테고리 추가"""
    try:
        data = await request.json()
        major = data.get("major")
        minor = data.get("minor")
        sub = data.get("sub")
        
        if not all([major, minor, sub]):
            return JSONResponse(status_code=400, content={"error": "대분류, 중분류, 소분류가 모두 필요합니다"})
        
        from ..services.effort_estimation import CategoryManager
        category_manager = CategoryManager()
        category_manager.add_category(major, minor, sub)
        
        return {"message": "카테고리가 추가되었습니다"}
    except Exception as e:
        logger.error(f"❌ 카테고리 추가 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.put("/effort/categories/")
async def update_category(request: Request):
    """카테고리 수정"""
    try:
        data = await request.json()
        old_major = data.get("old_major")
        old_minor = data.get("old_minor")
        old_sub = data.get("old_sub")
        new_major = data.get("new_major")
        new_minor = data.get("new_minor")
        new_sub = data.get("new_sub")
        
        if not all([old_major, old_minor, old_sub, new_major, new_minor, new_sub]):
            return JSONResponse(status_code=400, content={"error": "모든 필드가 필요합니다"})
        
        from ..services.effort_estimation import CategoryManager
        category_manager = CategoryManager()
        category_manager.update_category(old_major, old_minor, old_sub, new_major, new_minor, new_sub)
        
        return {"message": "카테고리가 수정되었습니다"}
    except Exception as e:
        logger.error(f"❌ 카테고리 수정 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.delete("/effort/categories/")
async def delete_category(request: Request):
    """카테고리 삭제"""
    try:
        data = await request.json()
        major = data.get("major")
        minor = data.get("minor")
        sub = data.get("sub")
        
        if not all([major, minor, sub]):
            return JSONResponse(status_code=400, content={"error": "대분류, 중분류, 소분류가 모두 필요합니다"})
        
        from ..services.effort_estimation import CategoryManager
        category_manager = CategoryManager()
        category_manager.delete_category(major, minor, sub)
        
        return {"message": "카테고리가 삭제되었습니다"}
    except Exception as e:
        logger.error(f"❌ 카테고리 삭제 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.put("/effort/update-category/")
async def update_effort_category(request: Request):
    """공수 산정 데이터의 카테고리 수정"""
    try:
        data = await request.json()
        jira_ticket = data.get("jira_ticket")
        major_category = data.get("major_category")
        minor_category = data.get("minor_category")
        sub_category = data.get("sub_category")
        
        if not all([jira_ticket, major_category, minor_category, sub_category]):
            return JSONResponse(status_code=400, content={"error": "모든 필드가 필요합니다"})
        
        from ..services.effort_estimation import effort_manager
        success = effort_manager.update_estimation_category(jira_ticket, major_category, minor_category, sub_category)
        
        if success:
            return {"message": "카테고리가 수정되었습니다"}
        else:
            return JSONResponse(status_code=404, content={"error": "해당 티켓을 찾을 수 없습니다"})
    except Exception as e:
        logger.error(f"❌ 카테고리 수정 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.delete("/effort/delete/{jira_ticket}")
async def delete_effort_estimation(jira_ticket: str):
    """공수 산정 데이터 삭제 (비활성화됨 - 데이터 보호)"""
    # 데이터 보호를 위해 삭제 기능 비활성화
    logger.warning(f"⚠️ 삭제 시도 차단: {jira_ticket} (삭제 기능 비활성화됨)")
    return JSONResponse(
        status_code=403, 
        content={
            "error": "데이터 보호를 위해 삭제 기능이 비활성화되었습니다",
            "message": "잘못된 데이터는 수정 기능을 사용하거나 관리자에게 문의하세요"
        }
    )
    
    # 원본 코드 (필요시 주석 해제)
    # try:
    #     from ..services.effort_estimation import effort_manager
    #     
    #     # 해당 티켓이 존재하는지 확인
    #     estimation = effort_manager.get_estimation_by_ticket(jira_ticket)
    #     if not estimation:
    #         return JSONResponse(status_code=404, content={"error": "해당 티켓을 찾을 수 없습니다"})
    #     
    #     # 삭제 실행
    #     success = effort_manager.delete_estimation(jira_ticket)
    #     
    #     if success:
    #         logger.info(f"✅ 공수 산정 데이터 삭제 완료: {jira_ticket}")
    #         return {"message": "데이터가 삭제되었습니다"}
    #     else:
    #         logger.error(f"❌ 공수 산정 데이터 삭제 실패: {jira_ticket}")
    #         return JSONResponse(status_code=500, content={"error": "삭제 중 오류가 발생했습니다"})
    # except Exception as e:
    #     logger.error(f"❌ 공수 산정 데이터 삭제 오류: {str(e)}")
    #     return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/effort/sync-epic/")
async def sync_epic_data(request: Request):
    """Epic 하위 작업 동기화"""
    try:
        # 요청 데이터 로깅
        logger.info(f"🔄 Epic 동기화 요청 수신 시작")
        
        # 동기화 전 데이터 백업
        from ..services.effort_estimation import effort_manager
        logger.info("💾 동기화 시작 전 데이터 백업 중...")
        effort_manager.backup_data()
        
        # Content-Type 확인
        content_type = request.headers.get("content-type", "")
        logger.info(f"🔄 Content-Type: {content_type}")
        
        # FormData 파싱
        form_data = await request.form()
        logger.info(f"🔄 FormData keys: {list(form_data.keys())}")
        
        epic_key = form_data.get("epic_key")
        major_category = form_data.get("major_category")
        minor_category = form_data.get("minor_category")
        sub_category = form_data.get("sub_category")
        title_filter = form_data.get("title_filter", "").strip()
        logger.info(f"🔄 Epic 동기화 요청 수신: epic_key={epic_key}, categories={major_category}/{minor_category}/{sub_category}, title_filter={title_filter}")
        
        if not epic_key:
            logger.error("❌ epic_key 파라미터가 없습니다")
            return JSONResponse(status_code=422, content={"error": "epic_key 파라미터가 필요합니다"})
        
        jira = create_jira_integration()
        if not jira:
            logger.error("❌ Jira 설정이 없습니다")
            return JSONResponse(status_code=400, content={"error": "Jira 설정이 필요합니다"})
        
        # Epic 기본 정보 조회 (Epic 이름 가져오기)
        epic_info = jira.test_epic_basic_info(epic_key)
        epic_name = "알 수 없음"
        if epic_info and isinstance(epic_info, dict):
            fields = epic_info.get('fields', {})
            epic_name = fields.get('summary', epic_key) if fields else epic_key
        logger.info(f"🔄 Epic 정보: {epic_key} - {epic_name}")
        
        # Epic 하위 작업 조회
        subtasks_result = jira.test_epic_subtasks(epic_key)
        if not subtasks_result or not subtasks_result.get("success"):
            error_msg = subtasks_result.get("error", "알 수 없는 오류") if subtasks_result else "Epic 조회 실패"
            logger.error(f"❌ Epic 하위 작업 조회 실패: {epic_key} - {error_msg}")
            return JSONResponse(status_code=404, content={"error": f"Epic '{epic_key}'의 하위 작업을 찾을 수 없습니다: {error_msg}"})
        
        # 작업 타입 필터링 (Epic만 제외하고 모든 타입 허용)
        tasks = subtasks_result.get("subtasks", [])
        excluded_types = ['Epic', '에픽']  # Epic 자체만 제외
        filtered_tasks = [task for task in tasks if task.get("issue_type") not in excluded_types]
        
        logger.info(f"🔄 작업 타입 필터링: 총 {len(tasks)}개 → {len(filtered_tasks)}개")
        logger.info(f"🔄 실제 타입들: {[task.get('issue_type') for task in tasks[:10]]}")  # 처음 10개만 로깅
        logger.info(f"🔄 제외된 타입: {excluded_types}")
        
        # 제목 필터링 (선택사항)
        if title_filter:
            original_count = len(filtered_tasks)
            filtered_tasks = [task for task in filtered_tasks if title_filter.lower() in task.get("summary", "").lower()]
            logger.info(f"🔄 제목 필터링: '{title_filter}' - {original_count}개 → {len(filtered_tasks)}개")
        
        logger.info(f"🔄 Epic '{epic_key}' 하위 작업: 총 {len(tasks)}개, 필터링 후 {len(filtered_tasks)}개")
        
        if not filtered_tasks:
            return JSONResponse(status_code=404, content={"error": f"Epic '{epic_key}'에 하위 작업이 없거나 모두 Epic 타입입니다"})
        
        # 각 작업을 공수 산정 데이터로 변환
        from ..services.effort_estimation import effort_manager
        
        added_count = 0
        updated_count = 0
        skipped_count = 0
        
        for task in filtered_tasks:
            try:
                # 기존 데이터 확인
                existing = effort_manager.get_estimation_by_ticket(task["key"])
                
                if existing:
                    # 기존 데이터 업데이트 (카테고리)
                    effort_manager.update_estimation_category(
                        task["key"], 
                        major_category or existing.major_category or "",
                        minor_category or existing.minor_category or "",
                        sub_category or existing.sub_category or ""
                    )
                    # Epic 정보 업데이트
                    effort_manager.update_estimation_epic(
                        task["key"],
                        epic_key,
                        epic_name
                    )
                    updated_count += 1
                    logger.info(f"✅ 기존 데이터 업데이트: {task['key']} (Epic: {epic_key})")
                else:
                    # 새 데이터 추가 (description 포함, comments만 제외)
                    from ..services.effort_estimation import EffortEstimation
                    
                    new_estimation = EffortEstimation(
                        jira_ticket=task["key"],
                        title=task["summary"],
                        story_points=task.get("story_points", 0),
                        description=task.get("description", None),  # description 포함
                        comments=None,  # comments만 제외
                        team_member=task.get("assignee", ""),
                        estimation_reason="Epic 하위 작업 자동 동기화",
                        major_category=major_category or "",
                        minor_category=minor_category or "",
                        sub_category=sub_category or "",
                        epic_key=epic_key,
                        epic_name=epic_name,
                        story_points_original=task.get("story_points_original"),
                        story_points_unit=task.get("story_points_unit", "M/D")
                    )
                    
                    effort_manager.add_estimation(new_estimation)
                    added_count += 1
                    logger.info(f"✅ 새 데이터 추가: {task['key']} (Epic: {epic_key})")
                    
            except Exception as e:
                logger.error(f"❌ 작업 처리 실패 {task['key']}: {str(e)}")
                skipped_count += 1
        
        # Epic 동기화 완료 후 증분 색인은 별도 배치로 실행 (속도 개선)
        # if added_count > 0 or updated_count > 0:
        #     logger.info("🔄 Epic 동기화 후 증분 색인 시작")
        #     try:
        #         # 추가/수정된 티켓 목록 수집
        #         synced_tickets = [task["key"] for task in filtered_tasks]
        #         
        #         if synced_tickets:
        #             json_file_path = os.path.join(DOCS_DIR, "effort_estimations.json")
        #             index_json_data_incremental(synced_tickets, json_file_path)
        #             logger.info(f"✅ Epic 동기화 후 증분 색인 완료: {len(synced_tickets)}개 티켓")
        #     except Exception as reindex_error:
        #         logger.warning(f"⚠️ Epic 동기화 후 증분 색인 실패 (무시하고 계속): {str(reindex_error)}")
        
        result = {
            "success": True,
            "epic_key": epic_key,
            "total_tasks": len(filtered_tasks),
            "added_tasks": added_count,
            "updated_tasks": updated_count,
            "skipped_tasks": skipped_count,
            "jql_used": subtasks_result.get("jql_used", "알 수 없음"),
            "message": f"Epic '{epic_key}' 하위 작업 동기화 완료 (색인은 '데이터 재색인' 버튼으로 별도 실행)"
        }
        
        logger.info(f"✅ Epic 동기화 완료: {result}")
        return result
        
    except Exception as e:
        logger.error(f"❌ Epic 동기화 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

def save_scheduler_history(scheduler_name: str, status: str, details: dict, start_time=None, end_time=None):
    """스케줄러 실행 이력 저장 (성공/실패만 기록, 실행중은 제외)"""
    try:
        # "running" 상태는 저장하지 않음
        if status == "running":
            return
        
        history_file = os.path.join(DOCS_DIR, "scheduler_history.json")
        
        # 기존 이력 로드
        history = []
        if os.path.exists(history_file):
            with open(history_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
        
        # 새 이력 추가
        history.append({
            "scheduler_name": scheduler_name,
            "status": status,  # "success" or "failed"
            "start_time": start_time.isoformat() if start_time else datetime.now().isoformat(),
            "end_time": end_time.isoformat() if end_time else datetime.now().isoformat(),
            "details": details
        })
        
        # 최근 100개만 유지
        history = history[-100:]
        
        # 저장
        with open(history_file, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        
        logger.info(f"✅ 스케줄러 이력 저장: {scheduler_name} - {status}")
        
    except Exception as e:
        logger.error(f"❌ 스케줄러 이력 저장 오류: {str(e)}")

def sync_completed_epics_background():
    """완료된 Epic 자동 동기화 백그라운드 작업 (ENOMIX 프로젝트만)"""
    global sync_status
    
    start_time = datetime.now()
    
    try:
        logger.info(f"🔄 완료된 Epic 자동 동기화 백그라운드 작업 시작 (ENOMIX 프로젝트)")
        
        # 동기화 전 데이터 백업
        from ..services.effort_estimation import effort_manager
        logger.info("💾 동기화 시작 전 데이터 백업 중...")
        effort_manager.backup_data()
        
        # 상태 초기화
        sync_status["is_running"] = True
        sync_status["progress"] = 0
        sync_status["completed_epics"] = 0
        sync_status["failed_epics"] = 0
        sync_status["current_epic"] = ""
        sync_status["message"] = f"Jira에서 완료된 Epic 검색 중 (ENOMIX 프로젝트)..."
        sync_status["failed_list"] = []
        
        jira = create_jira_integration()
        if not jira:
            sync_status["is_running"] = False
            sync_status["message"] = "Jira 설정이 없습니다"
            logger.error("❌ Jira 설정이 없습니다")
            return
        
        # ENOMIX는 기본 필드만 (빠름)
        include_details = False
        
        # 1. 완료된 Epic 목록 조회 (ENOMIX만)
        completed_epics = jira.search_completed_epics()
        
        if not completed_epics:
            sync_status["is_running"] = False
            sync_status["message"] = "완료된 Epic이 없습니다"
            sync_status["progress"] = 100
            logger.info("ℹ️ 완료된 Epic이 없습니다")
            return
        
        sync_status["total_epics"] = len(completed_epics)
        sync_status["message"] = f"{len(completed_epics)}개 Epic 동기화 시작"
        logger.info(f"🔍 완료된 Epic {len(completed_epics)}개 발견")
        
        # 2. 각 Epic 동기화
        for idx, epic in enumerate(completed_epics, 1):
            epic_key = epic['key']
            
            try:
                sync_status["current_epic"] = f"{epic_key} - {epic['summary'][:30]}..."
                sync_status["message"] = f"동기화 중: {epic_key} ({idx}/{len(completed_epics)})"
                logger.info(f"🔄 Epic 동기화 중: {epic_key} - {epic['summary'][:50]}...")
                
                # Epic 정보 조회
                epic_info = jira.test_epic_basic_info(epic_key)
                epic_name = "알 수 없음"
                if epic_info and isinstance(epic_info, dict):
                    fields = epic_info.get('fields', {})
                    epic_name = fields.get('summary', epic_key) if fields else epic_key
                
                # Epic 하위 작업 조회 (프로젝트별로 상세 정보 포함 여부 결정)
                subtasks_result = jira.test_epic_subtasks(epic_key, include_details=include_details)
                if not subtasks_result or not subtasks_result.get("success"):
                    logger.warning(f"⚠️ Epic {epic_key} 하위 작업 없음")
                    sync_status["failed_epics"] += 1
                    sync_status["failed_list"].append(f"{epic_key} (하위 작업 없음)")
                    continue
                
                # 작업 타입 필터링 (Epic만 제외하고 모든 타입 허용)
                tasks = subtasks_result.get("subtasks", [])
                excluded_types = ['Epic', '에픽']  # Epic 자체만 제외
                filtered_tasks = [task for task in tasks if task.get("issue_type") not in excluded_types]
                
                if not filtered_tasks:
                    logger.warning(f"⚠️ Epic {epic_key} 하위 작업 없음 (Epic 타입만 있음)")
                    sync_status["failed_epics"] += 1
                    sync_status["failed_list"].append(f"{epic_key} (하위 작업 없음)")
                    continue
                
                # 각 작업을 공수 산정 데이터로 변환
                from ..services.effort_estimation import effort_manager, EffortEstimation
                
                task_added = 0
                task_updated = 0
                
                for task in filtered_tasks:
                    try:
                        existing = effort_manager.get_estimation_by_ticket(task["key"])
                        
                        if existing:
                            # 기존 데이터 Epic 정보 업데이트
                            effort_manager.update_estimation_epic(task["key"], epic_key, epic_name)
                            task_updated += 1
                        else:
                            # 새 데이터 추가 (description 포함, comments만 제외)
                            new_estimation = EffortEstimation(
                                jira_ticket=task["key"],
                                title=task["summary"],
                                story_points=task.get("story_points", 0),
                                description=task.get("description", None),  # description 포함
                                comments=None,  # comments만 제외
                                team_member=task.get("assignee", ""),
                                estimation_reason="완료된 Epic 자동 동기화",
                                major_category="",
                                minor_category="",
                                sub_category="",
                                epic_key=epic_key,
                                epic_name=epic_name,
                                story_points_original=task.get("story_points_original"),
                                story_points_unit=task.get("story_points_unit", "M/D")
                            )
                            effort_manager.add_estimation(new_estimation)
                            task_added += 1
                            
                    except Exception as task_error:
                        logger.error(f"❌ Task {task['key']} 처리 실패: {str(task_error)}")
                        continue
                
                logger.info(f"✅ Epic {epic_key} 동기화 완료: {task_added}개 추가, {task_updated}개 업데이트")
                
                # 증분 색인은 별도 배치 작업으로 실행 (속도 개선)
                # if task_added > 0 or task_updated > 0:
                #     try:
                #         synced_tickets = [task["key"] for task in filtered_tasks]
                #         json_file_path = os.path.join(DOCS_DIR, "effort_estimations.json")
                #         index_json_data_incremental(synced_tickets, json_file_path)
                #         logger.info(f"   ✅ 증분 색인 완료: {len(synced_tickets)}개 티켓")
                #     except Exception as index_error:
                #         logger.warning(f"   ⚠️ 증분 색인 실패 (무시하고 계속): {str(index_error)}")
                
                sync_status["completed_epics"] += 1
                
            except Exception as epic_error:
                logger.error(f"❌ Epic {epic_key} 동기화 실패: {str(epic_error)}")
                sync_status["failed_epics"] += 1
                sync_status["failed_list"].append(f"{epic_key} ({str(epic_error)})")
                continue
            
            # 진행률 업데이트
            sync_status["progress"] = int((idx / len(completed_epics)) * 100)
        
        # 완료 - 색인은 별도 배치 작업으로 실행
        sync_status["is_running"] = False
        sync_status["progress"] = 100
        sync_status["current_epic"] = ""
        sync_status["message"] = f"동기화 완료: {sync_status['completed_epics']}개 성공, {sync_status['failed_epics']}개 실패 (색인은 별도 실행 필요)"
        logger.info(f"✅ 완료된 Epic 자동 동기화 완료: {sync_status['message']}")
        logger.info(f"💡 벡터 DB 색인은 '데이터 재색인' 버튼으로 별도 실행하세요")
        
        # 성공 이력 저장
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        save_scheduler_history(
            "Epic 자동 동기화",
            "success",
            {
                "total_epics": sync_status["total_epics"],
                "completed_epics": sync_status["completed_epics"],
                "failed_epics": sync_status["failed_epics"],
                "failed_list": sync_status["failed_list"],
                "duration_seconds": duration,
                "message": sync_status["message"]
            },
            start_time=start_time,
            end_time=end_time
        )
        
    except Exception as e:
        logger.error(f"❌ 완료된 Epic 자동 동기화 오류: {str(e)}")
        sync_status["is_running"] = False
        sync_status["message"] = f"오류 발생: {str(e)}"
        
        # 실패 이력 저장
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        save_scheduler_history(
            "Epic 자동 동기화",
            "failed",
            {
                "error": str(e),
                "duration_seconds": duration,
                "message": sync_status["message"]
            },
            start_time=start_time,
            end_time=end_time
        )

@app.post("/effort/auto-sync-completed-epics/")
async def auto_sync_completed_epics(background_tasks: BackgroundTasks):
    """완료된 Epic 자동 동기화 시작 (백그라운드, ENOMIX 프로젝트만)"""
    global sync_status
    
    # 이미 실행 중인지 확인
    if sync_status["is_running"]:
        return {
            "success": False,
            "message": "이미 동기화가 진행 중입니다",
            "is_running": True
        }
    
    # Jira 설정 확인
    jira = create_jira_integration()
    if not jira:
        return JSONResponse(status_code=400, content={"error": "Jira 설정이 필요합니다"})
    
    logger.info(f"🔄 완료된 Epic 자동 동기화 시작: ENOMIX 프로젝트")
    
    # 백그라운드 작업 등록
    background_tasks.add_task(sync_completed_epics_background)
    
    return {
        "success": True,
        "message": "완료된 Epic 자동 동기화가 시작되었습니다 (ENOMIX). 백그라운드에서 진행됩니다.",
        "is_running": True
    }

@app.get("/effort/sync-status/")
async def get_sync_status():
    """동기화 상태 조회"""
    return sync_status

@app.get("/effort/scheduler-history/")
async def get_scheduler_history():
    """스케줄러 실행 이력 조회"""
    try:
        history_file = os.path.join(DOCS_DIR, "scheduler_history.json")
        
        if not os.path.exists(history_file):
            return {
                "success": True,
                "history": [],
                "message": "스케줄러 실행 이력이 없습니다"
            }
        
        with open(history_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
        
        # 최신 순으로 정렬
        history.reverse()
        
        return {
            "success": True,
            "history": history,
            "total": len(history)
        }
        
    except Exception as e:
        logger.error(f"❌ 스케줄러 이력 조회 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

# ==================== 프롬프트 관리 엔드포인트 ====================

@app.get("/prompts/intent/stats/")
async def get_intent_prompt_stats():
    """의도 분류 프롬프트 통계 조회"""
    try:
        stats = intent_prompt_manager.get_stats()
        return stats
    except Exception as e:
        logger.error(f"❌ 프롬프트 통계 조회 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/prompts/intent/related/")
async def add_related_example(example: str = Form(...)):
    """관련 예시 추가"""
    try:
        success = intent_prompt_manager.add_related_example(example)
        if success:
            return {"message": f"관련 예시가 추가되었습니다: {example}"}
        else:
            return {"message": f"이미 존재하는 예시입니다: {example}"}
    except Exception as e:
        logger.error(f"❌ 관련 예시 추가 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/prompts/intent/unrelated/")
async def add_unrelated_example(example: str = Form(...)):
    """관련 없는 예시 추가"""
    try:
        success = intent_prompt_manager.add_unrelated_example(example)
        if success:
            return {"message": f"관련 없는 예시가 추가되었습니다: {example}"}
        else:
            return {"message": f"이미 존재하는 예시입니다: {example}"}
    except Exception as e:
        logger.error(f"❌ 관련 없는 예시 추가 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.delete("/prompts/intent/related/")
async def remove_related_example(example: str = Form(...)):
    """관련 예시 제거"""
    try:
        success = intent_prompt_manager.remove_related_example(example)
        if success:
            return {"message": f"관련 예시가 제거되었습니다: {example}"}
        else:
            return {"message": f"존재하지 않는 예시입니다: {example}"}
    except Exception as e:
        logger.error(f"❌ 관련 예시 제거 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.delete("/prompts/intent/unrelated/")
async def remove_unrelated_example(example: str = Form(...)):
    """관련 없는 예시 제거"""
    try:
        success = intent_prompt_manager.remove_unrelated_example(example)
        if success:
            return {"message": f"관련 없는 예시가 제거되었습니다: {example}"}
        else:
            return {"message": f"존재하지 않는 예시입니다: {example}"}
    except Exception as e:
        logger.error(f"❌ 관련 없는 예시 제거 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

# ==================== 피드백 수집 엔드포인트 ====================

@app.get("/test/simple")
async def test_simple():
    """간단한 테스트 엔드포인트"""
    return {"message": "테스트 성공", "status": "ok"}

@app.get("/test/epic-list")
async def test_epic_list():
    """사용 가능한 Epic 목록 조회"""
    try:
        jira = create_jira_integration()
        
        # JQL로 Epic 타입 이슈 조회 (API v3) - 페이징 처리로 더 많은 Epic 조회
        search_url = f"{jira.jira_url}/rest/api/3/search/jql"
        
        all_epics = []
        start_at = 0
        max_results = 100
        
        # 먼저 특정 프로젝트의 Epic 조회 시도
        project_epics = []
        try:
            params_project = {
                'jql': 'project = ENOMIX AND issuetype = Epic ORDER BY created ASC',
                'maxResults': 500,  # 200에서 500으로 증가
                'fields': 'key,summary,status,resolution,created'
            }
            response_project = jira.session.get(search_url, params=params_project)
            if response_project.status_code == 200:
                project_results = response_project.json()
                project_epics = project_results.get('issues', [])
                logger.info(f"🔍 프로젝트별 Epic 조회: {len(project_epics)}개")
        except Exception as e:
            logger.warning(f"⚠️ 프로젝트별 Epic 조회 실패: {e}")
        
        # 프로젝트별 조회가 성공하면 그것을 사용, 아니면 전체 조회
        if project_epics:
            all_epics = project_epics
        else:
            while True:
                # 날짜 조건 없이 모든 Epic 조회 (더 넓은 범위)
                params = {
                    'jql': 'issuetype = Epic ORDER BY created ASC',  # 오래된 것부터 조회
                    'maxResults': max_results,
                    'startAt': start_at,
                    'fields': 'key,summary,status,resolution,created'
                }
                
                response = jira.session.get(search_url, params=params)
                
                if response.status_code != 200:
                    break
                    
                results = response.json()
                issues = results.get('issues', [])
                
                if not issues:
                    break
                    
                all_epics.extend(issues)
                
                # 더 이상 가져올 데이터가 없으면 중단
                if len(issues) < max_results:
                    break
                    
                start_at += max_results
                
                # 최대 500개까지만 조회 (무한 루프 방지)
                if start_at >= 500:
                    break
        
        # 결과를 response 형태로 변환
        response_data = {
            'issues': all_epics,
            'total': len(all_epics)
        }
        
        # Epic 목록 처리
        epics = []
        for issue in response_data.get('issues', []):
            epics.append({
                'key': issue['key'],
                'summary': issue['fields']['summary'],
                'status': issue['fields']['status']['name'],
                'created': issue['fields'].get('created', 'N/A')
            })
        
        # ENOMIX-7338이 목록에 있는지 확인
        target_epic = next((epic for epic in epics if epic['key'] == 'ENOMIX-7338'), None)
        if target_epic:
            logger.info(f"✅ ENOMIX-7338 발견: {target_epic}")
        else:
            logger.warning(f"⚠️ ENOMIX-7338이 목록에 없음. 총 {len(epics)}개 Epic 조회됨")
            
            # ENOMIX-7338을 직접 검색해보기
            try:
                params_direct = {
                    'jql': 'key = ENOMIX-7338',
                    'maxResults': 1,
                    'fields': 'key,summary,status,issuetype,assignee,created'
                }
                response_direct = jira.session.get(search_url, params=params_direct)
                if response_direct.status_code == 200:
                    direct_results = response_direct.json()
                    if direct_results.get('total', 0) > 0:
                        direct_epic = direct_results['issues'][0]
                        logger.info(f"🔍 ENOMIX-7338 직접 검색 성공: {direct_epic['key']} - {direct_epic['fields']['summary']}")
                    else:
                        logger.warning(f"🔍 ENOMIX-7338 직접 검색 결과 없음")
                else:
                    logger.warning(f"🔍 ENOMIX-7338 직접 검색 실패: {response_direct.status_code}")
            except Exception as e:
                logger.warning(f"🔍 ENOMIX-7338 직접 검색 오류: {e}")
            
        return {
            "success": True,
            "epics": epics,
            "total": len(epics),
            "test_urls": [f"/test/epic-info/{epic['key']}" for epic in epics[:3]]  # 테스트용 URL 추가
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "message": "Epic 목록 조회 중 오류 발생"
        }

@app.get("/test/jira-connection")
async def test_jira_connection():
    """Jira 연결 테스트"""
    try:
        jira = create_jira_integration()
        connection_result = jira.test_connection()
        
        return {
            "success": connection_result,
            "message": "Jira 연결 성공" if connection_result else "Jira 연결 실패",
            "jira_url": jira.jira_url if hasattr(jira, 'jira_url') else 'N/A',
            "username": jira.username if hasattr(jira, 'username') else 'N/A'
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "message": "Jira 연결 테스트 중 오류 발생"
        }

@app.get("/test/issue-all-fields/{ticket_key}")
async def test_issue_all_fields(ticket_key: str):
    """티켓의 모든 필드 조회 (디버깅용)"""
    try:
        logger.info(f"🔍 티켓 전체 필드 조회: {ticket_key}")
        
        jira = create_jira_integration()
        if not jira:
            return JSONResponse(status_code=400, content={"error": "Jira 설정이 필요합니다"})
        
        # API v3로 모든 필드 조회
        url = f"{jira.jira_url}/rest/api/3/issue/{ticket_key}"
        
        logger.info(f"🔄 Jira API 호출: {url}")
        response = jira.session.get(url)  # 필드 제한 없음 (모든 필드)
        
        if response.status_code == 200:
            data = response.json()
            fields = data.get('fields', {})
            
            # customfield만 추출
            custom_fields = {}
            for key, value in fields.items():
                if key.startswith('customfield_'):
                    custom_fields[key] = {
                        'value': value,
                        'type': type(value).__name__
                    }
            
            return {
                "success": True,
                "ticket_key": ticket_key,
                "total_fields": len(fields),
                "total_custom_fields": len(custom_fields),
                "custom_fields": custom_fields,
                "all_fields": fields  # 모든 필드 포함
            }
        else:
            return JSONResponse(
                status_code=response.status_code, 
                content={"error": f"Jira API 오류: {response.text}"}
            )
            
    except Exception as e:
        logger.error(f"❌ 티켓 조회 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/test/epic-subtasks/{epic_key}")
async def test_epic_subtasks(epic_key: str):
    """Epic 하위 Task 조회 테스트"""
    try:
        logger.info(f"🔍 Epic 하위 Task 조회 시도: {epic_key}")
        
        jira = create_jira_integration()
        result = jira.test_epic_subtasks(epic_key)
        
        # 디버깅을 위한 추가 정보
        result["debug_info"] = {
            "epic_key": epic_key,
            "timestamp": str(datetime.now()),
            "jira_url": jira.jira_url if jira else "N/A"
        }
        
        return result
    except Exception as e:
        logger.error(f"❌ Epic 하위 Task 조회 API 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e), "details": "서버 내부 오류가 발생했습니다."})

@app.get("/test/jql/{jql_query}")
async def test_jql_query(jql_query: str):
    """JQL 쿼리 직접 테스트"""
    try:
        jira = create_jira_integration()
        if not jira:
            return JSONResponse(status_code=400, content={"error": "Jira 설정이 필요합니다"})
        
        search_url = f"{jira.jira_url}/rest/api/3/search/jql"
        
        headers = {
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }
        
        params = {
            'jql': jql_query,
            'maxResults': 10,
            'fields': 'key,summary,status,issuetype,assignee',
            'expand': 'changelog'
        }
        
        logger.info(f"🔍 JQL 테스트 요청 URL: {search_url}")
        logger.info(f"🔍 JQL 테스트 요청 파라미터: {params}")
        logger.info(f"🔍 JQL 테스트 요청 헤더: {headers}")
        
        response = jira.session.get(search_url, params=params, headers=headers)
        
        return {
            "success": response.status_code == 200,
            "status_code": response.status_code,
            "jql_query": jql_query,
            "response": response.json() if response.status_code == 200 else response.text
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/test/issue-id/{issue_id}")
async def test_issue_id(issue_id: str):
    """이슈 ID로 조회 테스트"""
    try:
        jira = create_jira_integration()
        if not jira:
            return JSONResponse(status_code=400, content={"error": "Jira 설정이 필요합니다"})
        
        search_url = f"{jira.jira_url}/rest/api/3/search/jql"
        
        headers = {
            'Accept': 'application/json',
            'Content-Type': 'application/json'
        }
        
        params = {
            'jql': f'id = {issue_id}',
            'maxResults': 10,
            'fields': 'key,summary,status,issuetype,assignee,id',
            'expand': 'changelog'
        }
        
        logger.info(f"🔍 이슈 ID 테스트 요청 URL: {search_url}")
        logger.info(f"🔍 이슈 ID 테스트 요청 파라미터: {params}")
        logger.info(f"🔍 이슈 ID 테스트 요청 헤더: {headers}")
        
        response = jira.session.get(search_url, params=params, headers=headers)
        
        return {
            "success": response.status_code == 200,
            "status_code": response.status_code,
            "issue_id": issue_id,
            "response": response.json() if response.status_code == 200 else response.text
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/test/permissions")
async def test_permissions():
    """현재 계정의 권한 확인"""
    try:
        jira = create_jira_integration()
        if not jira:
            return JSONResponse(status_code=400, content={"error": "Jira 설정이 필요합니다"})
        
        # 현재 사용자 정보 조회
        user_url = f"{jira.jira_url}/rest/api/3/myself"
        user_response = jira.session.get(user_url)
        
        # 프로젝트 목록 조회
        projects_url = f"{jira.jira_url}/rest/api/3/project"
        projects_response = jira.session.get(projects_url)
        
        # ENOMIX 프로젝트 상세 정보 조회
        enomix_url = f"{jira.jira_url}/rest/api/3/project/ENOMIX"
        enomix_response = jira.session.get(enomix_url)
        
        return {
            "success": True,
            "user_info": user_response.json() if user_response.status_code == 200 else f"사용자 정보 조회 실패: {user_response.status_code}",
            "projects": projects_response.json() if projects_response.status_code == 200 else f"프로젝트 목록 조회 실패: {projects_response.status_code}",
            "enomix_project": enomix_response.json() if enomix_response.status_code == 200 else f"ENOMIX 프로젝트 조회 실패: {enomix_response.status_code}",
            "user_status": user_response.status_code,
            "projects_status": projects_response.status_code,
            "enomix_status": enomix_response.status_code
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/test/epic-info/{epic_key}")
async def test_epic_info(epic_key: str):
    """Epic 기본 정보 조회 테스트"""
    try:
        logger.info(f"🔍 Epic 정보 조회 시도: {epic_key}")
        
        # Jira 연결 테스트
        jira = create_jira_integration()
        connection_result = jira.test_connection()
        logger.info(f"Jira 연결 결과: {connection_result}")
        
        if not connection_result:
            return {
                "success": False,
                "error": "Jira 연결 실패",
                "details": "Jira 서버에 연결할 수 없습니다. 인증 정보를 확인해주세요."
            }
        
        # Epic 정보 조회
        epic_info = jira.test_epic_basic_info(epic_key)
        logger.info(f"Epic 정보 조회 결과: {epic_info}")
        logger.info(f"Epic 정보 타입: {type(epic_info)}")
        
        if epic_info is None:
            return {
                "success": False,
                "error": "Epic 정보 조회 실패",
                "details": f"Epic '{epic_key}'를 찾을 수 없습니다. Epic 키를 확인해주세요.",
                "debug_info": "epic_info가 None으로 반환됨"
            }
        
        if not isinstance(epic_info, dict):
            return {
                "success": False,
                "error": "Epic 정보 형식 오류",
                "details": f"예상된 형식이 아닙니다. 받은 타입: {type(epic_info)}",
                "debug_info": str(epic_info)
            }
        
        if 'fields' not in epic_info:
            return {
                "success": False,
                "error": "Epic 정보 필드 누락",
                "details": "Epic 정보에 'fields' 필드가 없습니다.",
                "debug_info": str(epic_info)
            }
        
        # Epic 정보를 안전하게 처리
        try:
            fields = epic_info.get('fields', {}) if epic_info else {}
            epic_title = 'N/A'
            issue_type = 'N/A'
            status = 'N/A'
            assignee = 'N/A'
            
            if fields:
                epic_title = fields.get('summary', 'N/A') if fields.get('summary') is not None else 'N/A'
                
                issue_type_obj = fields.get('issuetype')
                if issue_type_obj and isinstance(issue_type_obj, dict):
                    issue_type = issue_type_obj.get('name', 'N/A')
                
                status_obj = fields.get('status')
                if status_obj and isinstance(status_obj, dict):
                    status = status_obj.get('name', 'N/A')
                
                assignee_obj = fields.get('assignee')
                if assignee_obj and isinstance(assignee_obj, dict):
                    assignee = assignee_obj.get('displayName', 'N/A')
            
            return {
                "success": True,
                "epic_key": epic_key,
                "epic_title": epic_title,
                "issue_type": issue_type,
                "status": status,
                "assignee": assignee
            }
        except Exception as e:
            logger.error(f"❌ Epic 정보 처리 오류: {str(e)}")
            return {
                "success": False,
                "error": "Epic 정보 처리 오류",
                "details": f"Epic 정보를 처리하는 중 오류가 발생했습니다: {str(e)}",
                "debug_info": str(epic_info)
            }
    except Exception as e:
        logger.error(f"❌ Epic 정보 조회 API 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e), "details": "서버 내부 오류가 발생했습니다."})

@app.get("/test/epic-full-details/{epic_key}")
async def test_epic_full_details(epic_key: str):
    """Epic의 모든 필드와 링크 정보 조회 (상세 디버깅용)"""
    try:
        jira = create_jira_integration()
        if not jira:
            return JSONResponse(status_code=400, content={"error": "Jira 설정이 필요합니다"})
        
        logger.info(f"🔍 Epic 상세 정보 조회: {epic_key}")
        
        # 1. Epic 자체의 모든 필드 조회
        issue_url = f"{jira.jira_url}/rest/api/3/issue/{epic_key}"
        params = {'expand': 'names,schema,operations,changelog'}
        
        epic_response = jira.session.get(issue_url, params=params)
        epic_data = epic_response.json() if epic_response.status_code == 200 else {"error": epic_response.text}
        
        # 2. Epic의 링크된 이슈들 조회
        links_url = f"{jira.jira_url}/rest/api/3/issue/{epic_key}?fields=issuelinks"
        links_response = jira.session.get(links_url)
        links_data = links_response.json() if links_response.status_code == 200 else {"error": links_response.text}
        
        # 3. Epic을 parent로 하는 하위 이슈 검색
        search_url = f"{jira.jira_url}/rest/api/3/search/jql"
        parent_jql = f'parent = {epic_key}'
        parent_params = {
            'jql': parent_jql,
            'maxResults': 50,
            'fields': 'key,summary,issuetype,parent'
        }
        parent_response = jira.session.get(search_url, params=parent_params)
        parent_data = parent_response.json() if parent_response.status_code == 200 else {"error": parent_response.text}
        
        # 4. Epic Link 필드로 연결된 이슈 검색
        epiclink_jql = f'"Epic Link" = {epic_key}'
        epiclink_params = {
            'jql': epiclink_jql,
            'maxResults': 50,
            'fields': 'key,summary,issuetype,customfield_10014,customfield_10015'
        }
        epiclink_response = jira.session.get(search_url, params=epiclink_params)
        epiclink_data = epiclink_response.json() if epiclink_response.status_code == 200 else {"error": epiclink_response.text}
        
        # 5. 모든 커스텀 필드 중 Epic 관련 필드 찾기
        fields_url = f"{jira.jira_url}/rest/api/3/field"
        fields_response = jira.session.get(fields_url)
        all_fields = fields_response.json() if fields_response.status_code == 200 else []
        
        epic_related_fields = []
        if isinstance(all_fields, list):
            for field in all_fields:
                field_name = field.get('name', '').lower()
                if 'epic' in field_name or 'parent' in field_name:
                    epic_related_fields.append({
                        'id': field.get('id'),
                        'name': field.get('name'),
                        'type': field.get('schema', {}).get('type', 'N/A')
                    })
        
        return {
            "success": True,
            "epic_key": epic_key,
            "epic_full_data": epic_data,
            "linked_issues": links_data,
            "parent_search_result": {
                "jql": parent_jql,
                "total": parent_data.get("total", 0) if isinstance(parent_data, dict) else 0,
                "issues": parent_data.get("issues", []) if isinstance(parent_data, dict) else []
            },
            "epiclink_search_result": {
                "jql": epiclink_jql,
                "total": epiclink_data.get("total", 0) if isinstance(epiclink_data, dict) else 0,
                "issues": epiclink_data.get("issues", []) if isinstance(epiclink_data, dict) else []
            },
            "epic_related_fields": epic_related_fields,
            "response_codes": {
                "epic": epic_response.status_code,
                "links": links_response.status_code,
                "parent_search": parent_response.status_code,
                "epiclink_search": epiclink_response.status_code,
                "fields": fields_response.status_code
            }
        }
        
    except Exception as e:
        logger.error(f"❌ Epic 상세 정보 조회 오류: {str(e)}")
        import traceback
        return JSONResponse(status_code=500, content={
            "error": str(e),
            "traceback": traceback.format_exc()
        })

@app.post("/feedback/")
async def collect_feedback(request: Request):
    """사용자 피드백 수집"""
    try:
        data = await request.json()
        question = data.get("question")
        feedback_type = data.get("feedback_type")
        timestamp = data.get("timestamp")
        
        logger.info(f"📝 피드백 수신: {feedback_type} - '{question}'")
        
        # 피드백에 따라 프롬프트 업데이트
        if feedback_type == "helpful":
            # 도움이 된 질문 → RELATED 예시에 추가
            success = intent_prompt_manager.add_related_example(question)
            if success:
                logger.info(f"✅ RELATED 예시 추가: '{question}'")
            else:
                logger.info(f"ℹ️ 이미 존재하는 RELATED 예시: '{question}'")
                
        elif feedback_type == "not-helpful":
            # 도움이 안된 질문 → UNRELATED 예시에 추가
            success = intent_prompt_manager.add_unrelated_example(question)
            if success:
                logger.info(f"✅ UNRELATED 예시 추가: '{question}'")
            else:
                logger.info(f"ℹ️ 이미 존재하는 UNRELATED 예시: '{question}'")
                
        elif feedback_type == "wrong-classification":
            # 잘못 분류된 질문 → 반대 카테고리로 이동
            related_examples = intent_prompt_manager.get_related_examples()
            unrelated_examples = intent_prompt_manager.get_unrelated_examples()
            
            if question in related_examples:
                # RELATED에서 UNRELATED로 이동
                intent_prompt_manager.remove_related_example(question)
                intent_prompt_manager.add_unrelated_example(question)
                logger.info(f"🔄 '{question}' RELATED → UNRELATED 이동")
            elif question in unrelated_examples:
                # UNRELATED에서 RELATED로 이동
                intent_prompt_manager.remove_unrelated_example(question)
                intent_prompt_manager.add_related_example(question)
                logger.info(f"🔄 '{question}' UNRELATED → RELATED 이동")
            else:
                # 새로운 질문이면 UNRELATED에 추가
                intent_prompt_manager.add_unrelated_example(question)
                logger.info(f"✅ 새로운 UNRELATED 예시 추가: '{question}'")
        
        return {"message": "피드백이 성공적으로 처리되었습니다"}
        
    except Exception as e:
        logger.error(f"❌ 피드백 처리 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/effort/categories/migrate/")
async def migrate_categories(request: Request):
    """카테고리 변경 시 기존 데이터 마이그레이션"""
    try:
        data = await request.json()
        old_category = data.get("old_category")  # "대분류 > 중분류 > 소분류"
        new_category = data.get("new_category")  # "대분류 > 중분류 > 소분류"
        
        if not old_category or not new_category:
            return JSONResponse(status_code=400, content={"error": "old_category와 new_category가 필요합니다"})
        
        # 카테고리 파싱
        old_parts = old_category.split(' > ')
        new_parts = new_category.split(' > ')
        
        if len(old_parts) != 3 or len(new_parts) != 3:
            return JSONResponse(status_code=400, content={"error": "카테고리는 대분류 > 중분류 > 소분류 형식이어야 합니다"})
        
        old_major, old_minor, old_sub = old_parts
        new_major, new_minor, new_sub = new_parts
        
        # 해당 카테고리의 모든 데이터 업데이트
        estimations = effort_manager.get_all_estimations()
        updated_count = 0
        
        for estimation in estimations:
            if (estimation.major_category == old_major and 
                estimation.minor_category == old_minor and 
                estimation.sub_category == old_sub):
                
                estimation.major_category = new_major
                estimation.minor_category = new_minor
                estimation.sub_category = new_sub
                updated_count += 1
        
        # 변경사항 저장
        effort_manager.save_data()
        
        return {
            "message": f"{updated_count}개 데이터가 마이그레이션되었습니다",
            "updated_count": updated_count
        }
        
    except Exception as e:
        logger.error(f"❌ 카테고리 마이그레이션 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/effort/categories/upload-excel")
async def upload_categories_excel(file: UploadFile = File(...)):
    """엑셀 파일로 카테고리 업데이트"""
    try:
        # 엑셀 파일 읽기 (openpyxl로 병합된 셀 처리)
        contents = await file.read()
        from openpyxl import load_workbook
        
        # openpyxl로 워크북 로드
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # JSON 구조로 변환
        categories = {}
        
        # 병합된 셀 정보 수집
        merged_ranges = list(ws.merged_cells.ranges)
        
        for row in ws.iter_rows(min_row=2, values_only=True):  # 헤더 제외
            if not any(row):  # 빈 행 건너뛰기
                continue
                
            # 병합된 셀 값 처리
            major = str(row[0]).strip() if row[0] and str(row[0]).strip() != 'None' else ""
            minor = str(row[1]).strip() if row[1] and str(row[1]).strip() != 'None' else ""
            sub = str(row[2]).strip() if row[2] and str(row[2]).strip() != 'None' else ""
            
            # 병합된 셀에서 값이 비어있으면 이전 행의 값 사용
            if not major and categories:
                # 이전 대분류 값 사용
                major = list(categories.keys())[-1] if categories else ""
            
            if not minor and major in categories and categories[major]:
                # 이전 중분류 값 사용
                minor = list(categories[major].keys())[-1] if categories[major] else ""
            
            if major and major != 'nan':
                if major not in categories:
                    categories[major] = {}
                if minor and minor != 'nan':
                    if minor not in categories[major]:
                        categories[major][minor] = []
                    if sub and sub != 'nan':
                        categories[major][minor].append(sub)
        
        # 기존 파일 백업
        categories_file = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'docs', 'categories.json')
        backup_file = f"{categories_file}.backup"
        if os.path.exists(categories_file):
            shutil.copy2(categories_file, backup_file)
        
        # 새 카테고리 저장
        with open(categories_file, 'w', encoding='utf-8') as f:
            json.dump(categories, f, ensure_ascii=False, indent=2)
        
        return {"success": True, "message": "엑셀 파일로 카테고리 업데이트 완료", "categories": categories}
    except Exception as e:
        logger.error(f"❌ 엑셀 파일 업로드 오류: {str(e)}")
        return {"success": False, "error": str(e)}

@app.get("/effort/categories/download-excel")
async def download_categories_excel():
    """현재 카테고리를 엑셀 파일로 다운로드"""
    try:
        # 현재 카테고리 로드
        categories_file = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'docs', 'categories.json')
        
        if not os.path.exists(categories_file):
            return JSONResponse(status_code=404, content={"error": "카테고리 파일을 찾을 수 없습니다"})
        
        with open(categories_file, 'r', encoding='utf-8') as f:
            categories = json.load(f)
        
        # openpyxl로 워크북 생성
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "카테고리"
        
        # 헤더 설정
        headers = ['대분류', '중분류', '소분류']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 데이터 입력 및 병합 처리
        current_row = 2
        major_start_row = 2
        minor_start_row = 2
        
        for major, minor_categories in categories.items():
            major_start_row = current_row
            
            for minor, sub_categories in minor_categories.items():
                minor_start_row = current_row
                
                for sub in sub_categories:
                    ws.cell(row=current_row, column=1, value=major)
                    ws.cell(row=current_row, column=2, value=minor)
                    ws.cell(row=current_row, column=3, value=sub)
                    current_row += 1
                
                # 중분류 병합 (소분류가 여러 개인 경우)
                if current_row - minor_start_row > 1:
                    ws.merge_cells(f'B{minor_start_row}:B{current_row - 1}')
            
            # 대분류 병합 (중분류가 여러 개인 경우)
            if current_row - major_start_row > 1:
                ws.merge_cells(f'A{major_start_row}:A{current_row - 1}')
        
        # 스타일 적용
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 모든 셀에 테두리 적용
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 메모리에서 파일 생성
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 파일명 생성 (현재 날짜/시간 포함)
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"categories_{timestamp}.xlsx"
        
        # 임시 파일로 저장 후 FileResponse 반환
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(output.getvalue())
            tmp_file_path = tmp_file.name
        
        return FileResponse(
            path=tmp_file_path,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"❌ 엑셀 파일 다운로드 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/effort/feedback/")
async def save_positive_feedback(request: Request):
    """피드백 데이터 저장 (긍정/부정 모두 지원)"""
    try:
        data = await request.json()
        question = data.get("question", "")
        answer = data.get("answer", "")
        sources = data.get("sources", [])
        feedback_type = data.get("feedback_type", "positive")
        user = data.get("user", "web")  # 웹에서 온 피드백은 "web"으로 표시
        
        logger.info(f"💾 피드백 저장 요청: {feedback_type} - {question[:50]}...")
        
        # 피드백 데이터 구성
        feedback_data = {
            "question": question,
            "answer": answer,
            "sources": sources,
            "timestamp": datetime.now().isoformat(),
            "feedback_type": feedback_type,
            "source": "web",  # 웹에서 온 피드백
            "user": user
        }
        
        # 피드백 저장 (긍정/부정 모두 저장)
        result = save_feedback_to_file(feedback_data)
        
        if result.get("saved"):
            feedback_count = result.get("feedback_count", 1)
            is_new = result.get("is_new", True)
            
            if is_new:
                logger.info(f"✅ 피드백 저장 완료 (새로운 세트): {feedback_type} - {question[:30]}...")
            else:
                logger.info(f"📊 피드백 카운트 증가: {feedback_type} - {question[:30]}... (총 {feedback_count}회)")
            
            return {
                "status": "success",
                "message": "피드백이 저장되었습니다",
                "feedback_count": feedback_count,
                "is_new": is_new
            }
        else:
            logger.error(f"❌ 피드백 저장 실패: {feedback_type} - {question[:30]}...")
            return JSONResponse(
                status_code=500,
                content={"error": "피드백 저장에 실패했습니다"}
            )
        
    except Exception as e:
        logger.error(f"❌ 피드백 저장 오류: {str(e)}")
        import traceback
        logger.error(f"❌ 상세 오류: {traceback.format_exc()}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/effort/reindex-json/")
async def reindex_json_data(background_tasks: BackgroundTasks):
    """JSON 파일 강제 재인덱싱 (백그라운드 실행)"""
    try:
        json_file_path = os.path.join(DOCS_DIR, "effort_estimations.json")
        if not os.path.exists(json_file_path):
            return JSONResponse(status_code=404, content={"error": "effort_estimations.json 파일을 찾을 수 없습니다"})
        
        # 백그라운드로 재인덱싱 실행
        background_tasks.add_task(reindex_json_background, json_file_path)
        
        logger.info("🔄 JSON 파일 재인덱싱 백그라운드 작업 시작")
        return {
            "status": "started", 
            "message": "재인덱싱이 백그라운드에서 시작되었습니다. 완료까지 수 분 소요될 수 있습니다."
        }
        
    except Exception as e:
        logger.error(f"❌ JSON 파일 재인덱싱 오류: {str(e)}")
        return JSONResponse(status_code=500, content={"error": str(e)})

def reindex_json_background(json_file_path: str):
    """재인덱싱 백그라운드 작업"""
    try:
        logger.info("📚 백그라운드 재인덱싱 시작...")
        start_time = time.time()
        
        result = index_json_data(json_file_path, force=True)
        
        elapsed = time.time() - start_time
        if result:
            logger.info(f"✅ 백그라운드 재인덱싱 완료 (소요 시간: {elapsed:.1f}초)")
        else:
            logger.error(f"❌ 백그라운드 재인덱싱 실패 (소요 시간: {elapsed:.1f}초)")
            
    except Exception as e:
        logger.error(f"❌ 백그라운드 재인덱싱 오류: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())

# StaticFiles 마운트 - API 라우트들 뒤에 배치
app.mount("/effort-management", StaticFiles(directory=os.path.join(STATIC_DIR, "effort-management")), name="effort-management")
app.mount("/category-management", StaticFiles(directory=os.path.join(STATIC_DIR, "category-management")), name="category-management")